import os
import sys
import argparse
import glob
import time
import datetime
import cv2
import numpy as np    
import pyautogui
from ultralytics import YOLO
from openpyxl import Workbook

#----------------- 🔔 Play Sound Notification for new detection----------

import platform

def play_beep_sound(path):
    if platform.system() == 'Windows':
        import winsound
        winsound.PlaySound(path, winsound.SND_FILENAME | winsound.SND_ASYNC)
    else:
        try:
            from playsound import playsound
            playsound(path, block=False)
        except Exception as e:
            print(f"[⚠️ SOUND ERROR] {e}")


# ------------------------- Argument Parsing -------------------------
parser = argparse.ArgumentParser()
parser.add_argument('--model', required=True, help='Path to YOLO model (.pt)')
parser.add_argument('--source', required=True, help='Input source: image/video/folder/cam index')
parser.add_argument('--thresh', type=float, default=0.5, help='Confidence threshold')
parser.add_argument('--resolution', default=None, help='Resolution WxH (e.g. 1280x720)')
parser.add_argument('--record', action='store_true', help='Record output to a video file')
args = parser.parse_args()

# ------------------------- Input Values -------------------------
model_path = args.model
img_source = args.source
min_thresh = args.thresh
user_res = args.resolution
record = args.record

# ------------------------- Validate Model -------------------------
if not os.path.exists(model_path):
    print(f"[❌ ERROR] Model not found: {model_path}")
    sys.exit(1)

# ------------------------- Load Model -------------------------
model = YOLO(model_path, task='detect')
labels = model.names

# ------------------------- Input Source Check -------------------------
img_ext_list = ['.jpg', '.jpeg', '.png', '.bmp']
vid_ext_list = ['.mp4', '.avi', '.mov', '.mkv', '.wmv']

if os.path.isdir(img_source):
    source_type = 'folder'
elif os.path.isfile(img_source):
    _, ext = os.path.splitext(img_source)
    if ext.lower() in img_ext_list:
        source_type = 'image'
    elif ext.lower() in vid_ext_list:
        source_type = 'video'
    else:
        print(f"[❌ ERROR] Unsupported file extension: {ext}")
        sys.exit(1)
elif img_source.isdigit():
    source_type = 'usb'
    usb_idx = int(img_source)
elif "http" in img_source and "8080" in img_source:
    source_type = 'ipcam'
else:
    print(f"[❌ ERROR] Invalid input source: {img_source}")
    sys.exit(1)

# ------------------------- Resolution -------------------------
resize = False
if user_res:
    resize = True
    try:
        resW, resH = map(int, user_res.lower().split('x'))
    except:
        print("[❌ ERROR] Invalid resolution format. Use WxH like 640x480")
        sys.exit(1)


# ✅ Initialize product counter early
product_counter = {}

# ------------------------- Setup Output Paths -------------------------
timestamp = datetime.datetime.now().strftime("%d_%m_%Y_%I.%M_%p_IST")
output_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Output_of_Estimation")
os.makedirs(output_dir, exist_ok=True)

record_path = os.path.join(output_dir, f"inference_{timestamp}.avi")
excel_path = os.path.join(output_dir, f"Packers_and_Movers_Estimation_{timestamp}.xlsx")
beep_sound_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "beep.wav")

# ------------------------- Predefined Charges -------------------------
charges_dict = {
    "Cooler": 200,
    "Furniture": 1500,
    "Iron_Box": 30,
    "Laptop": 150,
    "Microwave_Oven": 300,
    "Mixture": 100,
    "Phone": 50,
    "Refridgerator": 1200,
    "Remote": 10,
    "Standing_Fan": 100,
    "Tele_Vision": 500,
    "Vacuum_Cleaner": 100,
    "Washing_Machine": 1100
}

# ------------------------- Recording Setup -------------------------
if record:
    if source_type not in ['video', 'usb', 'ipcam']:
        print("[❌ ERROR] --record is only supported for video/camera sources.")
        sys.exit(1)
    if not user_res:
        print("[❌ ERROR] Please specify --resolution when recording.")
        sys.exit(1)
    recorder = cv2.VideoWriter(record_path, cv2.VideoWriter_fourcc(*'MJPG'), 30, (resW, resH))
    
# ------------------------- Input Capture -------------------------
if source_type == 'image':
    imgs_list = [img_source]

elif source_type == 'folder':
    imgs_list = [img for img in glob.glob(os.path.join(img_source, '*'))
                 if os.path.splitext(img)[1].lower() in img_ext_list]

elif source_type == 'video':
    cap = cv2.VideoCapture(img_source)
    if resize:
        cap.set(3, resW)
        cap.set(4, resH)

elif source_type == 'usb':
    cap = cv2.VideoCapture(usb_idx)
    if resize:
        cap.set(3, resW)
        cap.set(4, resH)

elif source_type == 'ipcam':
    cap = cv2.VideoCapture(img_source)
    if not cap.isOpened():
        print(f"[❌ ERROR] Failed to open IP camera stream at {img_source}")
    else:
        print(f"[✅] Connected to IP camera stream at {img_source}")
    if resize:
        cap.set(3, resW)
        cap.set(4, resH)

else:
    print(f"[⚠️] Unknown source_type: {source_type}")

# ------------------------- Inference Loop Setup -------------------------
bbox_colors = [(255, 0, 0), (0, 255, 0), (0, 0, 255),
               (255, 255, 0), (255, 0, 255), (0, 255, 255)]

fps_list = []
img_count = 0
screen_w, screen_h = pyautogui.size()

# ------------------------- Inference Loop -------------------------
while True:
    t_start = time.perf_counter()

    if source_type in ['image', 'folder']:
        if img_count >= len(imgs_list):
            print("[✅ DONE] All images processed.")
            break
        frame = cv2.imread(imgs_list[img_count])
        img_count += 1
    else:
        ret, frame = cap.read()
        if not ret or frame is None:
            print("[❌ ERROR] Frame not received from camera. Trying next frame...")
            continue

    if resize:
        frame = cv2.resize(frame, (resW, resH))

    results = model(frame, verbose=False)
    detections = results[0].boxes
    
    # Track per-frame class instance count
    frame_seen_labels = set()
    frame_instance_counts = {}

    for det in detections:
        conf = det.conf.item()
        if conf < min_thresh:
            continue

        xmin, ymin, xmax, ymax = map(int, det.xyxy.cpu().numpy().squeeze())
        classid = int(det.cls.item())
        label_name = labels[classid]

        frame_instance_counts[label_name] = frame_instance_counts.get(label_name, 0) + 1

        label = f"{label_name}: {int(conf * 100)}%"
        color = bbox_colors[classid % len(bbox_colors)]
        cv2.rectangle(frame, (xmin, ymin), (xmax, ymax), color, 2)
        cv2.putText(frame, label, (xmin, max(ymin - 10, 10)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, color, 2)

    # ✅ Define new_detections AFTER the loop
    new_detections = len(frame_instance_counts) > 0

    # ✅ Count unique new detections only
    for label in frame_instance_counts:
        if label not in product_counter:
            product_counter[label] = 1

    # 🔔 Play Sound Notification for new detection
    if new_detections and os.path.exists(beep_sound_path):
        play_beep_sound(beep_sound_path)

    # Display Product Counts
    line_height = 20
    start_y = 80
    for idx, prod in enumerate(sorted(product_counter)):
        display_text = f"{prod} : 01"
        y = start_y + idx * line_height
        cv2.putText(frame, display_text, (10, y), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 255, 0), 2)


    # Display FPS & Meta
    fps = 1 / (time.perf_counter() - t_start)
    fps_list.append(fps)
    if len(fps_list) > 100: fps_list.pop(0)
    avg_fps = sum(fps_list) / len(fps_list)

    cv2.putText(frame, f"FPS: {avg_fps:.2f}", (10, 40), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,255), 2)
    cv2.putText(frame, f"Objects: {len(frame_seen_labels)}", (10, 65), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0,255,255), 2)

    wm1, wm2 = "Model : Packers_and_Movers_Estimation", "Created By : Mohammed Gouse"
    font = cv2.FONT_HERSHEY_SIMPLEX
    scale = 0.6
    thickness = 2
    color = (255, 255, 255)

    (text_w1, _), _ = cv2.getTextSize(wm1, font, scale, thickness)
    (text_w2, _), _ = cv2.getTextSize(wm2, font, scale, thickness)

    x1 = frame.shape[1] - text_w1 - 10
    x2 = frame.shape[1] - text_w2 - 10
    y1 = 50   # ⬅ moved slightly down
    y2 = y1 + 25

    cv2.putText(frame, wm1, (x1, y1), font, scale, color, thickness)
    cv2.putText(frame, wm2, (x2, y2), font, scale, color, thickness)

    scale_ratio = min(screen_w / frame.shape[1], screen_h / frame.shape[0])
    display_frame = cv2.resize(frame, None, fx=scale_ratio, fy=scale_ratio)

    cv2.imshow("YOLOv8 Detection", display_frame)
    if record: recorder.write(frame)

    key = cv2.waitKey(5)
    if key in [ord('q'), ord('Q')]: break
    elif key in [ord('s'), ord('S')]:
        cv2.imwrite(os.path.join(output_dir, 'capture.jpg'), frame)

# ------------------------- Write to Excel -------------------------
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

if 'product_counter' not in locals():
    product_counter = {}

wb = Workbook()
ws = wb.active
ws.title = "Estimation"

# ------------------ Styles ------------------
header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
header_font = Font(bold=True)
bold_font = Font(bold=True)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

thin = Side(style='thin')
thick = Side(style='thick')
thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
thick_border = Border(left=thick, right=thick, top=thick, bottom=thick)

# ------------------ Add Header ------------------
headers = ["S.No", "Product", "Product Count", "Transition Charges (Per Count)", "Transition Charges (Each Category)"]
ws.append(headers)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

# ------------------ Add Data Rows ------------------
total_count = 0
total_cost = 0
for idx, prod in enumerate(product_counter, 1):
    rate = charges_dict.get(prod, 0)
    cost = rate * 1
    total_count += 1
    total_cost += cost
    ws.append([idx, prod, 1, rate, cost])

# ------------------ Add Total Row ------------------
ws.append(["Total Product Count", "", total_count, "Estimated Cost", f"Rs.{total_cost}/-"])
total_row_idx = ws.max_row
ws.merge_cells(f"A{total_row_idx}:B{total_row_idx}")
ws[f"A{total_row_idx}"].alignment = center_align

# ------------------ Autofit Columns ------------------
for col in ws.columns:
    max_length = max(len(str(cell.value)) if cell.value else 0 for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws.column_dimensions[col_letter].width = max_length + 4

# ------------------ Apply Borders ------------------
for row in range(1, total_row_idx + 1):
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)

        # Determine border style (outer border = thick, inner = thin)
        top = thick if row == 1 else thin
        bottom = thick if row == total_row_idx else thin
        left = thick if col == 1 else thin
        right = thick if col == 5 else thin

        # Special case: entire first and last row should be thick
        if row in [1, total_row_idx]:
            top = bottom = left = right = thick

        # Apply border
        border = Border(top=top, bottom=bottom, left=left, right=right)

        # Style header row
        if row == 1:
            cell.font = header_font
            cell.fill = header_fill

        # Style total row
        elif row == total_row_idx:
            cell.font = bold_font
            cell.fill = light_green_fill

        cell.alignment = center_align
        cell.border = border
# ------------------ Save Workbook ------------------
wb.save(excel_path)
print(f"[📊 EXCEL SAVED] {excel_path}")

# -------------🚀 Auto-send Excel via email -------------

import smtplib
from email.message import EmailMessage

def send_email_with_excel(excel_path, receiver_email):
    sender_email = "myselfmohammedgouse@gmail.com"
    sender_password = "euin jrtu qtiw zomy".replace(" ", "")  # Use App Password if 2FA enabled

    msg = EmailMessage()
    msg["Subject"] = f"Packers and Movers Estimation - {timestamp}"
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg.set_content(
    """\
Dear Customer,

Please find the attached Estimation for your request regarding the Transition of your Home/Office Goods.

Have a nice day!!

Best Regards,  
Packers and Movers  
Email: email@email.com
Phone No: 9876543210
"""
)

    with open(excel_path, "rb") as f:
        file_data = f.read()
        file_name = os.path.basename(excel_path)
    msg.add_attachment(file_data, maintype="application", subtype="octet-stream", filename=file_name)

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(sender_email, sender_password)
            smtp.send_message(msg)
        print(f"[📧 EMAIL SENT] Log file sent to {receiver_email}")
    except Exception as e:
        print(f"[❌ EMAIL ERROR] {e}")


#  Calling Existing Function to send Excel via email
send_email_with_excel(excel_path, receiver_email="azeematall@gmail.com")

# ------------------------- Cleanup -------------------------
if source_type in ['video', 'usb', 'ipcam']:
    cap.release()
if record:
    recorder.release()
cv2.destroyAllWindows()

